"""
deed_queue_export.py - Export a deed/price capture queue for missing-price transactions.

Usage:
    python deed_queue_export.py
    python deed_queue_export.py --county Hernando
    python deed_queue_export.py --type "Builder Purchase"
    python deed_queue_export.py --exclude-inventory-category scattered_legacy_lots
    python deed_queue_export.py --limit 250 --out output/deed_queue.xlsx
"""

import argparse
import json
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from config import DATABASE_URL, OUTPUT_DIR


_COUNTY_PORTAL_URLS = {
    'Bay': 'https://www.baycoclerk.com/public-Records/official-Record-search/',
    'Citrus': 'https://search.citrusclerk.org/LandmarkWeb',
    'Santa Rosa': 'https://santarosaclerk.com/courts/search-public-records/',
    'Walton': 'https://waltonclerk.com/recordsearch',
}

_DETAIL_COLUMNS = [
    'ID',
    'County',
    'Date',
    'Type',
    'Grantor',
    'Grantee',
    'Instrument',
    'Subdivision',
    'Phase',
    'Inventory Category',
    'Lots',
    'Acres',
    'Export Legal Description',
    'Recommended Search',
    'Search Query',
    'Search Portal',
    'Price',
    'Book Type',
    'Book',
    'Page',
    'Book/Page',
    'Instrument Number',
    'CFN',
    'Clerk File Number',
    'File Number',
    'Case Number',
    'Reference',
    'Doc Link Text',
    'Image Link',
    'Locator JSON',
    'Source File',
]

_DETAIL_WIDTHS = {
    'ID': 10,
    'County': 12,
    'Date': 12,
    'Type': 18,
    'Grantor': 36,
    'Grantee': 36,
    'Instrument': 12,
    'Subdivision': 34,
    'Phase': 12,
    'Inventory Category': 24,
    'Lots': 8,
    'Acres': 10,
    'Export Legal Description': 52,
    'Recommended Search': 20,
    'Search Query': 28,
    'Search Portal': 42,
    'Price': 12,
    'Book Type': 10,
    'Book': 12,
    'Page': 12,
    'Book/Page': 16,
    'Instrument Number': 18,
    'CFN': 16,
    'Clerk File Number': 18,
    'File Number': 18,
    'Case Number': 16,
    'Reference': 18,
    'Doc Link Text': 44,
    'Image Link': 18,
    'Locator JSON': 60,
    'Source File': 34,
}

_SUMMARY_WIDTHS = {
    'Metric': 30,
    'Value': 12,
    'County': 16,
    'Rows': 12,
    'Strategy': 24,
    'Portal': 42,
    'Inventory Category': 24,
}

_CENTER_COLUMNS = {
    'ID',
    'County',
    'Date',
    'Instrument',
    'Phase',
    'Lots',
    'Price',
    'Book Type',
}

_WRAP_COLUMNS = {
    'Grantor',
    'Grantee',
    'Subdivision',
    'Export Legal Description',
    'Search Query',
    'Search Portal',
    'Doc Link Text',
    'Locator JSON',
    'Source File',
}


def build_query(county: str | None = None, transaction_type: str = 'Builder Purchase',
                limit: int | None = None,
                inventory_categories: list[str] | None = None,
                exclude_inventory_categories: list[str] | None = None) -> tuple[str, list]:
    where = ['price IS NULL']
    params: list = []

    if transaction_type:
        where.append('type = %s')
        params.append(transaction_type)

    if county:
        where.append("REPLACE(UPPER(county), ' ', '') = REPLACE(UPPER(%s), ' ', '')")
        params.append(county)
    if inventory_categories:
        where.append('(' + ' OR '.join(['inventory_category = %s'] * len(inventory_categories)) + ')')
        params.extend(inventory_categories)
    if exclude_inventory_categories:
        where.extend(['inventory_category IS DISTINCT FROM %s'] * len(exclude_inventory_categories))
        params.extend(exclude_inventory_categories)

    sql = """
        SELECT
            id,
            county,
            date,
            grantor,
            grantee,
            type,
            instrument,
            subdivision,
            phase,
            inventory_category,
            lots,
            acres,
            export_legal_desc,
            price,
            source_file,
            deed_locator
        FROM transactions
    """
    if where:
        sql += ' WHERE ' + ' AND '.join(where)
    sql += ' ORDER BY county, date NULLS LAST, id'

    if limit is not None:
        sql += ' LIMIT %s'
        params.append(limit)

    return sql, params


def _coerce_json(value) -> dict:
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip():
        try:
            loaded = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return loaded if isinstance(loaded, dict) else {}
    return {}


def _list_text(values) -> str:
    if not values:
        return ''
    if not isinstance(values, list):
        values = [values]

    cleaned = []
    for value in values:
        text = str(value).strip()
        if text and text.lower() != 'nan' and text not in cleaned:
            cleaned.append(text)
    return ' | '.join(cleaned)


def _is_http_url(value: str | None) -> bool:
    if not value:
        return False
    parsed = urlparse(value)
    return parsed.scheme in {'http', 'https'} and bool(parsed.netloc)


def _book_page_text(locator: dict) -> str:
    book_page = str(locator.get('book_page') or '').strip()
    if book_page:
        return book_page

    book = str(locator.get('book') or '').strip()
    page = str(locator.get('page') or '').strip()
    if book and page:
        return f'{book}/{page}'
    return ''


def recommend_search(locator: dict) -> str:
    doc_link = str(locator.get('doc_link') or '').strip()
    if _is_http_url(doc_link):
        return 'Direct URL'
    if locator.get('instrument_number'):
        return 'Instrument Number'
    if locator.get('cfn'):
        return 'CFN'
    if locator.get('file_number'):
        return 'File Number'
    if locator.get('book') and locator.get('page'):
        return 'Book/Page'
    if locator.get('clerk_file_number'):
        return 'Clerk File Number'
    if locator.get('reference'):
        return 'Reference'
    if locator.get('image_link'):
        return 'Image Reference'
    return 'Missing Locator'


def build_search_query(locator: dict) -> str:
    doc_link = str(locator.get('doc_link') or '').strip()
    if _is_http_url(doc_link):
        return doc_link
    if locator.get('instrument_number'):
        return f"Instrument {locator['instrument_number']}"
    if locator.get('cfn'):
        return f"CFN {locator['cfn']}"
    if locator.get('file_number'):
        return f"File No. {locator['file_number']}"
    if locator.get('book') and locator.get('page'):
        book_type = str(locator.get('book_type') or '').strip()
        prefix = f'{book_type} ' if book_type else ''
        return f"{prefix}{locator['book']}/{locator['page']}"
    if locator.get('clerk_file_number'):
        return f"Clerk File {locator['clerk_file_number']}"
    if locator.get('reference'):
        return f"Reference {locator['reference']}"
    if locator.get('image_link'):
        return f"Image {locator['image_link']}"
    return ''


def flatten_deed_row(row: dict) -> dict:
    locator = _coerce_json(row.get('deed_locator'))
    search_strategy = recommend_search(locator)
    doc_link = str(locator.get('doc_link') or '').strip()

    return {
        'ID': row.get('id'),
        'County': row.get('county'),
        'Date': row.get('date'),
        'Type': row.get('type'),
        'Grantor': row.get('grantor'),
        'Grantee': row.get('grantee'),
        'Instrument': row.get('instrument'),
        'Subdivision': row.get('subdivision'),
        'Phase': row.get('phase'),
        'Inventory Category': row.get('inventory_category'),
        'Lots': row.get('lots'),
        'Acres': row.get('acres'),
        'Export Legal Description': row.get('export_legal_desc'),
        'Recommended Search': search_strategy,
        'Search Query': build_search_query(locator),
        'Search Portal': _COUNTY_PORTAL_URLS.get(row.get('county') or '', ''),
        'Price': row.get('price'),
        'Book Type': locator.get('book_type'),
        'Book': locator.get('book'),
        'Page': locator.get('page'),
        'Book/Page': _book_page_text(locator),
        'Instrument Number': locator.get('instrument_number'),
        'CFN': locator.get('cfn'),
        'Clerk File Number': locator.get('clerk_file_number'),
        'File Number': locator.get('file_number'),
        'Case Number': locator.get('case_number'),
        'Reference': locator.get('reference'),
        'Doc Link Text': '' if _is_http_url(doc_link) else _list_text(locator.get('doc_links') or doc_link),
        'Image Link': locator.get('image_link'),
        'Locator JSON': json.dumps(locator, sort_keys=True),
        'Source File': row.get('source_file'),
    }


def fetch_queue_rows(sql: str, params: list) -> pd.DataFrame:
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            columns = [desc[0] for desc in cur.description]
            raw_df = pd.DataFrame(cur.fetchall(), columns=columns)
    finally:
        conn.close()

    rows = [flatten_deed_row(record) for record in raw_df.to_dict(orient='records')]
    return pd.DataFrame(rows, columns=_DETAIL_COLUMNS)


def build_summary_frames(detail_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    overview = pd.DataFrame([
        {'Metric': 'Queue Rows', 'Value': int(len(detail_df))},
        {'Metric': 'Counties Represented', 'Value': int(detail_df['County'].dropna().nunique())},
        {'Metric': 'Rows With Locator', 'Value': int((detail_df['Recommended Search'] != 'Missing Locator').sum())},
        {'Metric': 'Rows With Portal URL', 'Value': int(detail_df['Search Portal'].replace('', pd.NA).dropna().shape[0])},
    ])

    county_counts = (
        detail_df.groupby('County', dropna=False)
        .size()
        .rename('Rows')
        .reset_index()
        .sort_values(['Rows', 'County'], ascending=[False, True], kind='stable')
    )

    strategy_counts = (
        detail_df.groupby('Recommended Search', dropna=False)
        .size()
        .rename('Rows')
        .reset_index()
        .rename(columns={'Recommended Search': 'Strategy'})
        .sort_values(['Rows', 'Strategy'], ascending=[False, True], kind='stable')
    )
    category_counts = (
        detail_df.assign(
            **{'Inventory Category': detail_df['Inventory Category'].replace('', pd.NA).fillna('Uncategorized')}
        )
        .groupby('Inventory Category', dropna=False)
        .size()
        .rename('Rows')
        .reset_index()
        .sort_values(['Rows', 'Inventory Category'], ascending=[False, True], kind='stable')
    )

    return overview, county_counts, strategy_counts, category_counts


def _style_sheet(path: Path, sheet_name: str, widths: dict[str, int], *,
                 freeze_panes: str | None = None) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]

    if freeze_panes:
        ws.freeze_panes = freeze_panes

    for cell in ws[1]:
        cell.font = Font(bold=True)

    header_map = {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }

    for name, width in widths.items():
        col_idx = header_map.get(name)
        if col_idx:
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    for name in _CENTER_COLUMNS:
        col_idx = header_map.get(name)
        if not col_idx:
            continue
        for row_idx in range(1, ws.max_row + 1):
            ws.cell(row_idx, col_idx).alignment = Alignment(horizontal='center')

    for name in _WRAP_COLUMNS:
        col_idx = header_map.get(name)
        if not col_idx:
            continue
        for row_idx in range(1, ws.max_row + 1):
            ws.cell(row_idx, col_idx).alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(path)


def export_queue(detail_df: pd.DataFrame, out_path: Path) -> None:
    overview_df, county_counts_df, strategy_counts_df, category_counts_df = build_summary_frames(detail_df)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine='openpyxl', datetime_format='m/d/yyyy') as writer:
        detail_df.to_excel(writer, sheet_name='Deed Queue', index=False)
        overview_df.to_excel(writer, sheet_name='Summary', index=False, startrow=0)
        county_counts_df.to_excel(writer, sheet_name='Summary', index=False, startrow=len(overview_df) + 3)
        category_counts_df.to_excel(writer, sheet_name='Summary', index=False, startrow=len(overview_df) + 3, startcol=3)
        strategy_counts_df.to_excel(
            writer,
            sheet_name='Summary',
            index=False,
            startrow=len(overview_df) + max(len(county_counts_df), len(category_counts_df)) + 6,
        )

    _style_sheet(out_path, 'Deed Queue', _DETAIL_WIDTHS, freeze_panes='A2')
    _style_sheet(out_path, 'Summary', _SUMMARY_WIDTHS)


def main() -> None:
    parser = argparse.ArgumentParser(description='Export missing-price deed queue to Excel')
    parser.add_argument('--county', help='Filter by county name (spacing-insensitive)')
    parser.add_argument('--type', dest='transaction_type', default='Builder Purchase',
                        help='Transaction type filter (default: Builder Purchase)')
    parser.add_argument('--inventory-category', action='append',
                        help='Only export rows in the given inventory category (repeat for multiple)')
    parser.add_argument('--exclude-inventory-category', action='append',
                        help='Exclude rows in the given inventory category (repeat for multiple)')
    parser.add_argument('--limit', type=int, help='Optional max number of rows to export')
    parser.add_argument('--out', help='Output file path (default: OUTPUT_DIR/deed_queue.xlsx)')
    args = parser.parse_args()

    sql, params = build_query(
        args.county,
        args.transaction_type,
        args.limit,
        inventory_categories=args.inventory_category,
        exclude_inventory_categories=args.exclude_inventory_category,
    )
    detail_df = fetch_queue_rows(sql, params)

    if detail_df.empty:
        print('No records matched the filter criteria.')
        return

    out_path = Path(args.out) if args.out else OUTPUT_DIR / 'deed_queue.xlsx'
    export_queue(detail_df, out_path)
    print(f"Exported {len(detail_df):,} records -> {out_path}")


if __name__ == '__main__':
    main()
