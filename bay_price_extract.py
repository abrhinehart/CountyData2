"""
bay_price_extract.py - Extract Bay builder-purchase prices from Bay County official records.

Usage:
    python bay_price_extract.py
    python bay_price_extract.py --limit 5
    python bay_price_extract.py --apply
    python bay_price_extract.py --out output/bay_price_extract.xlsx
"""

import argparse
import json
import re
import time
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
import psycopg2
import psycopg2.extras
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from config import DATABASE_URL, OUTPUT_DIR


_BAY_HOME_URL = 'https://records2.baycoclerk.com/Recording/'
_RESULT_COLUMNS = [
    'Transaction ID',
    'Status',
    'Grantor',
    'Grantee',
    'Subdivision',
    'Phase',
    'Target Clerk File #',
    'Target Book/Page',
    'Matched Clerk File #',
    'Matched Book/Page',
    'Matched Record Date',
    'Matched Doc Type',
    'Matched Document ID',
    'Extracted Consideration',
    'Applied',
    'Note',
    'Source File',
]
_COLUMN_WIDTHS = {
    'Transaction ID': 12,
    'Status': 18,
    'Grantor': 36,
    'Grantee': 36,
    'Subdivision': 28,
    'Phase': 10,
    'Target Clerk File #': 18,
    'Target Book/Page': 16,
    'Matched Clerk File #': 18,
    'Matched Book/Page': 18,
    'Matched Record Date': 20,
    'Matched Doc Type': 14,
    'Matched Document ID': 18,
    'Extracted Consideration': 20,
    'Applied': 10,
    'Note': 34,
    'Source File': 34,
}


def fetch_bay_queue(limit: int | None = None) -> list[dict]:
    sql = """
        SELECT
            id,
            grantor,
            grantee,
            subdivision,
            phase,
            source_file,
            deed_locator,
            parsed_data
        FROM transactions
        WHERE county = 'Bay'
          AND type = 'Builder Purchase'
          AND price IS NULL
        ORDER BY date NULLS LAST, id
    """
    params: list = []
    if limit is not None:
        sql += ' LIMIT %s'
        params.append(limit)

    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(sql, params)
            return list(cur.fetchall())
    finally:
        conn.close()


def build_target_book_page(locator: dict) -> str:
    book = str(locator.get('book') or '').strip()
    page = str(locator.get('page') or '').strip()
    if book and page:
        return f'{book}/{page}'
    return ''


def _clean_html_text(value: str) -> str:
    text = re.sub(r'<br\s*/?>', '\n', value, flags=re.IGNORECASE)
    text = re.sub(r'<[^>]+>', ' ', text)
    text = text.replace('&nbsp;', ' ')
    text = text.replace('&amp;', '&')
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def extract_detail_field(detail_html: str, label: str) -> str | None:
    pattern = re.compile(
        rf'<label[^>]*>\s*{re.escape(label)}\s*</label>\s*</td>\s*<td[^>]*>\s*(.*?)<br',
        re.IGNORECASE | re.DOTALL,
    )
    match = pattern.search(detail_html)
    if not match:
        return None
    value = _clean_html_text(match.group(1))
    return value or None


def parse_bay_detail_html(detail_html: str) -> dict:
    return {
        'instrument_number': extract_detail_field(detail_html, 'Instrument #'),
        'book_page': extract_detail_field(detail_html, 'Book/Page'),
        'record_date': extract_detail_field(detail_html, 'Record Date'),
        'book_type': extract_detail_field(detail_html, 'Book Type'),
        'consideration_raw': extract_detail_field(detail_html, 'Consideration'),
    }


def parse_currency(value: str | None) -> Decimal | None:
    if not value:
        return None
    cleaned = re.sub(r'[^0-9.-]', '', value)
    if not cleaned:
        return None
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def parse_result_row(row_element) -> dict:
    from selenium.webdriver.common.by import By

    cells = row_element.find_elements(By.TAG_NAME, 'td')
    text = [cell.text.strip() for cell in cells]
    row_id = row_element.get_attribute('id') or ''
    document_id = None
    row_number = None
    match = re.match(r'doc_(\d+)_(\d+)', row_id)
    if match:
        document_id = match.group(1)
        row_number = int(match.group(2))

    return {
        'row_id': row_id,
        'row_number': row_number,
        'document_id': document_id,
        'grantor': text[5] if len(text) > 5 else '',
        'grantee': text[6] if len(text) > 6 else '',
        'record_date': text[7] if len(text) > 7 else '',
        'doc_type': text[8] if len(text) > 8 else '',
        'book_type': text[9] if len(text) > 9 else '',
        'book': text[10] if len(text) > 10 else '',
        'page': text[11] if len(text) > 11 else '',
        'clerk_file_number': text[12] if len(text) > 12 else '',
        'legal': text[13] if len(text) > 13 else '',
    }


def _find_exact_result(rows: list[dict], locator: dict) -> dict | None:
    clerk_file_number = str(locator.get('clerk_file_number') or '').strip()
    book = str(locator.get('book') or '').strip()
    page = str(locator.get('page') or '').strip().lstrip('0')

    for row in rows:
        if clerk_file_number and row.get('clerk_file_number') == clerk_file_number:
            return row

    for row in rows:
        row_book = str(row.get('book') or '').strip()
        row_page = str(row.get('page') or '').strip().lstrip('0')
        if book and page and row_book == book and row_page == page:
            return row

    return None


def _build_driver():
    from selenium import webdriver
    from selenium.webdriver.chrome.service import Service
    from webdriver_manager.chrome import ChromeDriverManager

    options = webdriver.ChromeOptions()
    options.add_argument('--window-size=1600,1200')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_experimental_option('excludeSwitches', ['enable-automation'])
    options.add_experimental_option('useAutomationExtension', False)

    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.execute_cdp_cmd(
        'Page.addScriptToEvaluateOnNewDocument',
        {'source': "Object.defineProperty(navigator, 'webdriver', {get: () => undefined});"},
    )
    return driver


def _open_bay_search(driver, wait) -> None:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC

    driver.get(_BAY_HOME_URL)
    time.sleep(2)
    driver.execute_script(
        "document.getElementById('goToSection').value='searchCriteriaInstrumentNumber'; SetDisclaimer();"
    )
    wait.until(EC.presence_of_element_located((By.ID, 'instrumentNumber')))
    # Give the page's recaptcha polling a moment to settle.
    time.sleep(6)


def _search_bay(driver, wait, clerk_file_number: str) -> list[dict]:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC

    _open_bay_search(driver, wait)
    field = wait.until(EC.element_to_be_clickable((By.ID, 'instrumentNumber')))
    field.clear()
    field.send_keys(clerk_file_number)
    driver.find_element(By.ID, 'submit-InstrumentNumber').click()
    wait.until(EC.presence_of_element_located((By.ID, 'resultsTable')))
    time.sleep(5)
    row_elements = driver.find_elements(By.CSS_SELECTOR, '#resultsTable tbody tr')
    return [parse_result_row(row) for row in row_elements]


def _fetch_detail(driver, wait, row: dict) -> dict:
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC

    row_element = wait.until(EC.presence_of_element_located((By.ID, row['row_id'])))
    driver.execute_script('arguments[0].scrollIntoView({block: "center"});', row_element)
    row_element.click()
    wait.until(lambda d: 'Consideration' in d.find_element(By.ID, 'detailSection').get_attribute('innerHTML'))
    time.sleep(2)
    detail_html = driver.find_element(By.ID, 'detailSection').get_attribute('innerHTML')
    detail = parse_bay_detail_html(detail_html)
    detail['document_id'] = row.get('document_id')
    return detail


def _build_price_extraction_payload(target: dict, matched_row: dict, detail: dict, price: Decimal) -> dict:
    return {
        'source': 'bay_clerk_detail',
        'method': 'selenium_visible_browser',
        'searched_by': 'clerk_file_number',
        'matched_document_id': matched_row.get('document_id'),
        'matched_clerk_file_number': matched_row.get('clerk_file_number'),
        'matched_book_page': detail.get('book_page') or f"{matched_row.get('book_type')} {matched_row.get('book')} / {matched_row.get('page')}",
        'matched_record_date': detail.get('record_date') or matched_row.get('record_date'),
        'raw_consideration': detail.get('consideration_raw'),
        'price': str(price),
        'extracted_at_utc': datetime.now(timezone.utc).isoformat(),
        'deed_locator': target.get('deed_locator') or {},
    }


def apply_result_to_db(transaction_id: int, price: Decimal, extraction_payload: dict) -> bool:
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE transactions
                SET price = %s,
                    parsed_data = jsonb_set(
                        COALESCE(parsed_data, '{}'::jsonb),
                        '{price_extraction}',
                        %s::jsonb,
                        TRUE
                    ),
                    updated_at = NOW()
                WHERE id = %s
                  AND price IS NULL
                """,
                (price, json.dumps(extraction_payload), transaction_id),
            )
            updated = cur.rowcount > 0
        conn.commit()
        return updated
    finally:
        conn.close()


def process_bay_queue(rows: list[dict], *, apply_updates: bool = False) -> list[dict]:
    from selenium.webdriver.support.ui import WebDriverWait

    driver = _build_driver()
    wait = WebDriverWait(driver, 45)
    results = []

    try:
        for row in rows:
            locator = row.get('deed_locator') or {}
            clerk_file_number = str(locator.get('clerk_file_number') or '').strip()
            target_book_page = build_target_book_page(locator)

            result = {
                'Transaction ID': row.get('id'),
                'Status': 'locator_missing',
                'Grantor': row.get('grantor'),
                'Grantee': row.get('grantee'),
                'Subdivision': row.get('subdivision'),
                'Phase': row.get('phase'),
                'Target Clerk File #': clerk_file_number,
                'Target Book/Page': target_book_page,
                'Matched Clerk File #': '',
                'Matched Book/Page': '',
                'Matched Record Date': '',
                'Matched Doc Type': '',
                'Matched Document ID': '',
                'Extracted Consideration': '',
                'Applied': 'No',
                'Note': '',
                'Source File': row.get('source_file'),
            }

            if not clerk_file_number:
                result['Note'] = 'Bay locator is missing clerk_file_number.'
                results.append(result)
                continue

            try:
                search_rows = _search_bay(driver, wait, clerk_file_number)
                matched_row = _find_exact_result(search_rows, locator)
                if not matched_row:
                    result['Status'] = 'no_exact_result'
                    result['Note'] = 'No exact Bay result matched the stored clerk file number or book/page.'
                    results.append(result)
                    continue

                detail = _fetch_detail(driver, wait, matched_row)
                price = parse_currency(detail.get('consideration_raw'))

                result['Matched Clerk File #'] = matched_row.get('clerk_file_number') or ''
                result['Matched Book/Page'] = detail.get('book_page') or f"{matched_row.get('book_type')} {matched_row.get('book')} / {matched_row.get('page')}"
                result['Matched Record Date'] = detail.get('record_date') or matched_row.get('record_date') or ''
                result['Matched Doc Type'] = matched_row.get('doc_type') or ''
                result['Matched Document ID'] = matched_row.get('document_id') or ''
                result['Extracted Consideration'] = detail.get('consideration_raw') or ''

                if price is None:
                    result['Status'] = 'no_consideration'
                    result['Note'] = 'Bay detail page did not expose a parseable consideration value.'
                    results.append(result)
                    continue

                result['Status'] = 'matched'
                if apply_updates:
                    extraction_payload = _build_price_extraction_payload(row, matched_row, detail, price)
                    applied = apply_result_to_db(row['id'], price, extraction_payload)
                    result['Applied'] = 'Yes' if applied else 'No'
                    if applied:
                        result['Note'] = 'Price applied to transactions.price and parsed_data.price_extraction.'
                    else:
                        result['Note'] = 'Price was already present or row could not be updated.'
                else:
                    result['Note'] = 'Dry run only. Re-run with --apply to write the extracted price.'

                results.append(result)
            except Exception as exc:  # noqa: BLE001 - we want row-level resilience
                result['Status'] = 'error'
                result['Note'] = f'{type(exc).__name__}: {exc}'
                results.append(result)
    finally:
        driver.quit()

    return results


def export_results(results: list[dict], out_path: Path) -> None:
    df = pd.DataFrame(results, columns=_RESULT_COLUMNS)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Bay Price Extract', index=False)

    wb = load_workbook(out_path)
    ws = wb['Bay Price Extract']
    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        cell.font = Font(bold=True)

    header_map = {
        str(ws.cell(1, c).value).strip(): c
        for c in range(1, ws.max_column + 1)
        if ws.cell(1, c).value
    }

    for name, width in _COLUMN_WIDTHS.items():
        col_idx = header_map.get(name)
        if col_idx:
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    for name in {'Transaction ID', 'Phase', 'Applied'}:
        col_idx = header_map.get(name)
        if not col_idx:
            continue
        for row_idx in range(1, ws.max_row + 1):
            ws.cell(row_idx, col_idx).alignment = Alignment(horizontal='center')

    for name in {'Grantor', 'Grantee', 'Subdivision', 'Note', 'Source File'}:
        col_idx = header_map.get(name)
        if not col_idx:
            continue
        for row_idx in range(1, ws.max_row + 1):
            ws.cell(row_idx, col_idx).alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description='Extract Bay builder-purchase prices from Bay County official records')
    parser.add_argument('--limit', type=int, help='Optional number of Bay rows to process')
    parser.add_argument('--apply', action='store_true', help='Write matched prices back to the database')
    parser.add_argument('--out', help='Output workbook path (default: OUTPUT_DIR/bay_price_extract.xlsx)')
    args = parser.parse_args()

    queue_rows = fetch_bay_queue(args.limit)
    if not queue_rows:
        print('No Bay Builder Purchase rows with missing price were found.')
        return

    results = process_bay_queue(queue_rows, apply_updates=args.apply)
    out_path = Path(args.out) if args.out else OUTPUT_DIR / 'bay_price_extract.xlsx'
    export_results(results, out_path)

    matched = sum(1 for row in results if row['Status'] == 'matched')
    applied = sum(1 for row in results if row['Applied'] == 'Yes')
    print(f"Processed {len(results)} Bay rows: {matched} matched, {applied} applied -> {out_path}")


if __name__ == '__main__':
    main()
