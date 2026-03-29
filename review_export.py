"""
review_export.py - Export the review queue to a workbook for triage.

Usage:
    python review_export.py
    python review_export.py --county Marion
    python review_export.py --reason subdivision_unmatched
    python review_export.py --reason multiple_subdivision_candidates --reason phase_not_confirmed_by_lookup
    python review_export.py --exclude-inventory-category scattered_legacy_lots
    python review_export.py --limit 250 --out review_queue.xlsx
"""

import argparse
import json
from pathlib import Path

import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from config import DATABASE_URL, OUTPUT_DIR


_DETAIL_COLUMNS = [
    'ID',
    'County',
    'Date',
    'Review Reasons',
    'Grantor',
    'Grantee',
    'Type',
    'Instrument',
    'Price',
    'Lots',
    'Inventory Category',
    'Subdivision',
    'Subdivision ID',
    'Phase',
    'Phase Candidates',
    'Lookup Text',
    'Preparsed Subdivision',
    'Ignored Subdivision Reason',
    'Normalized Candidates',
    'Lot Values',
    'Block Values',
    'Unit Values',
    'Parcel References',
    'Section Values',
    'Township Values',
    'Range Values',
    'Tract Values',
    'Subdivision Flags',
    'Export Legal Description',
    'Export Legal Raw',
    'Source File',
]

_DETAIL_WIDTHS = {
    'ID': 10,
    'County': 12,
    'Date': 12,
    'Review Reasons': 32,
    'Grantor': 36,
    'Grantee': 36,
    'Type': 20,
    'Instrument': 12,
    'Price': 12,
    'Lots': 8,
    'Inventory Category': 24,
    'Subdivision': 34,
    'Subdivision ID': 14,
    'Phase': 12,
    'Phase Candidates': 18,
    'Lookup Text': 34,
    'Preparsed Subdivision': 34,
    'Ignored Subdivision Reason': 24,
    'Normalized Candidates': 42,
    'Lot Values': 18,
    'Block Values': 18,
    'Unit Values': 18,
    'Parcel References': 24,
    'Section Values': 18,
    'Township Values': 18,
    'Range Values': 18,
    'Tract Values': 18,
    'Subdivision Flags': 24,
    'Export Legal Description': 64,
    'Export Legal Raw': 64,
    'Source File': 34,
}

_SUMMARY_WIDTHS = {
    'Metric': 28,
    'Value': 12,
    'Reason': 36,
    'Rows': 12,
    'County': 16,
    'Inventory Category': 24,
}

_CENTER_COLUMNS = {
    'ID',
    'County',
    'Date',
    'Instrument',
    'Price',
    'Lots',
    'Subdivision ID',
    'Phase',
}

_WRAP_COLUMNS = {
    'Review Reasons',
    'Lookup Text',
    'Preparsed Subdivision',
    'Normalized Candidates',
    'Export Legal Description',
    'Export Legal Raw',
    'Source File',
}


def build_query(county: str | None = None, reasons: list[str] | None = None,
                limit: int | None = None,
                inventory_categories: list[str] | None = None,
                exclude_inventory_categories: list[str] | None = None) -> tuple[str, list]:
    where = ['review_flag = TRUE']
    params: list = []

    if county:
        where.append("REPLACE(UPPER(county), ' ', '') = REPLACE(UPPER(%s), ' ', '')")
        params.append(county)

    if reasons:
        clauses = ["(parsed_data->'review_reasons') ? %s" for _ in reasons]
        where.append('(' + ' OR '.join(clauses) + ')')
        params.extend(reasons)
    if inventory_categories:
        where.append('(' + ' OR '.join(['inventory_category = %s'] * len(inventory_categories)) + ')')
        params.extend(inventory_categories)
    if exclude_inventory_categories:
        where.extend(['inventory_category IS DISTINCT FROM %s'] * len(exclude_inventory_categories))
        params.extend(exclude_inventory_categories)

    sql = """
        SELECT
            id,
            county,
            date,
            grantor,
            grantee,
            type,
            instrument,
            price,
            lots,
            inventory_category,
            subdivision,
            subdivision_id,
            phase,
            export_legal_desc,
            export_legal_raw,
            source_file,
            parsed_data
        FROM transactions
    """
    if where:
        sql += ' WHERE ' + ' AND '.join(where)
    sql += ' ORDER BY county, date NULLS LAST, id'

    if limit is not None:
        sql += ' LIMIT %s'
        params.append(limit)

    return sql, params


def _coerce_parsed_data(value) -> dict:
    if isinstance(value, dict):
        return value
    if isinstance(value, str) and value.strip():
        try:
            loaded = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return loaded if isinstance(loaded, dict) else {}
    return {}


def _list_text(values) -> str:
    if not values:
        return ''
    if not isinstance(values, list):
        values = [values]

    cleaned = []
    for value in values:
        text = str(value).strip()
        if text and text not in cleaned:
            cleaned.append(text)
    return ' | '.join(cleaned)


def _candidate_text(candidates: list[dict]) -> str:
    parts = []
    for candidate in candidates or []:
        subdivision = str(candidate.get('subdivision') or candidate.get('raw') or '').strip()
        if not subdivision:
            continue

        extras = []
        phase = str(candidate.get('phase') or '').strip()
        if phase:
            extras.append(f'phase={phase}')

        details = candidate.get('details') or {}
        alias_source = str(details.get('alias_source') or '').strip()
        if alias_source:
            extras.append(f'alias={alias_source}')

        text = subdivision
        if extras:
            text += f" ({', '.join(extras)})"
        if text not in parts:
            parts.append(text)

    return ' | '.join(parts)


def _county_parse_values(county_parse: dict, *keys: str) -> str:
    for key in keys:
        values = county_parse.get(key)
        text = _list_text(values)
        if text:
            return text
    return ''


def flatten_review_row(row: dict) -> dict:
    parsed_data = _coerce_parsed_data(row.get('parsed_data'))
    county_parse = parsed_data.get('county_parse') or {}
    review_reasons = parsed_data.get('review_reasons') or []

    return {
        'ID': row.get('id'),
        'County': row.get('county'),
        'Date': row.get('date'),
        'Review Reasons': _list_text(review_reasons),
        'Grantor': row.get('grantor'),
        'Grantee': row.get('grantee'),
        'Type': row.get('type'),
        'Instrument': row.get('instrument'),
        'Price': row.get('price'),
        'Lots': row.get('lots'),
        'Inventory Category': row.get('inventory_category'),
        'Subdivision': row.get('subdivision'),
        'Subdivision ID': row.get('subdivision_id'),
        'Phase': row.get('phase'),
        'Phase Candidates': _list_text(parsed_data.get('phase_candidate_values')),
        'Lookup Text': parsed_data.get('subdivision_lookup_text'),
        'Preparsed Subdivision': parsed_data.get('preparsed_subdivision'),
        'Ignored Subdivision Reason': parsed_data.get('ignored_subdivision_reason'),
        'Normalized Candidates': _candidate_text(county_parse.get('normalized_subdivision_candidates') or []),
        'Lot Values': _county_parse_values(county_parse, 'structured_lot_values', 'lot_values'),
        'Block Values': _county_parse_values(county_parse, 'structured_block_values', 'block_values'),
        'Unit Values': _county_parse_values(county_parse, 'structured_unit_values', 'unit_values', 'subdivision_unit_values'),
        'Parcel References': _county_parse_values(county_parse, 'structured_parcel_references', 'parcel_references'),
        'Section Values': _county_parse_values(county_parse, 'structured_section_values', 'section_values'),
        'Township Values': _county_parse_values(county_parse, 'structured_township_values', 'township_values'),
        'Range Values': _county_parse_values(county_parse, 'structured_range_values', 'range_values'),
        'Tract Values': _county_parse_values(county_parse, 'tract_values'),
        'Subdivision Flags': _county_parse_values(county_parse, 'subdivision_flags'),
        'Export Legal Description': row.get('export_legal_desc'),
        'Export Legal Raw': row.get('export_legal_raw'),
        'Source File': row.get('source_file'),
    }


def fetch_review_rows(sql: str, params: list) -> pd.DataFrame:
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            columns = [desc[0] for desc in cur.description]
            raw_df = pd.DataFrame(cur.fetchall(), columns=columns)
    finally:
        conn.close()

    rows = [flatten_review_row(record) for record in raw_df.to_dict(orient='records')]
    detail_df = pd.DataFrame(rows, columns=_DETAIL_COLUMNS)
    return detail_df


def build_summary_frames(detail_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    reason_series = (
        detail_df['Review Reasons']
        .replace('', pd.NA)
        .dropna()
        .str.split(r'\s*\|\s*')
        .explode()
        .dropna()
    )

    overview = pd.DataFrame([
        {'Metric': 'Flagged Rows', 'Value': int(len(detail_df))},
        {'Metric': 'Counties Represented', 'Value': int(detail_df['County'].dropna().nunique())},
        {'Metric': 'Distinct Review Reasons', 'Value': int(reason_series.nunique())},
    ])
    reason_counts = (
        reason_series.value_counts()
        .rename_axis('Reason')
        .reset_index(name='Rows')
    )

    county_counts = (
        detail_df.groupby('County', dropna=False)
        .size()
        .rename('Rows')
        .reset_index()
        .rename(columns={'county': 'County'})
        .sort_values(['Rows', 'County'], ascending=[False, True], kind='stable')
    )
    category_counts = (
        detail_df.assign(
            **{'Inventory Category': detail_df['Inventory Category'].replace('', pd.NA).fillna('Uncategorized')}
        )
        .groupby('Inventory Category', dropna=False)
        .size()
        .rename('Rows')
        .reset_index()
        .sort_values(['Rows', 'Inventory Category'], ascending=[False, True], kind='stable')
    )

    return overview, reason_counts, county_counts, category_counts


def _style_sheet(path: Path, sheet_name: str, widths: dict[str, int], *,
                 freeze_panes: str | None = None) -> None:
    wb = load_workbook(path)
    ws = wb[sheet_name]

    if freeze_panes:
        ws.freeze_panes = freeze_panes

    for cell in ws[1]:
        cell.font = Font(bold=True)

    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    headers = {
        str(ws.cell(1, col).value).strip(): col
        for col in range(1, ws.max_column + 1)
        if ws.cell(1, col).value
    }

    for header, width in widths.items():
        if header in headers:
            ws.column_dimensions[get_column_letter(headers[header])].width = width

    for header in _CENTER_COLUMNS:
        if header not in headers:
            continue
        col_index = headers[header]
        for row_index in range(1, ws.max_row + 1):
            ws.cell(row_index, col_index).alignment = Alignment(horizontal='center', vertical='top')

    for header in _WRAP_COLUMNS:
        if header not in headers:
            continue
        col_index = headers[header]
        for row_index in range(1, ws.max_row + 1):
            ws.cell(row_index, col_index).alignment = Alignment(wrap_text=True, vertical='top')

    if 'Date' in headers:
        col_index = headers['Date']
        for row_index in range(2, ws.max_row + 1):
            cell = ws.cell(row_index, col_index)
            if cell.value is not None:
                cell.number_format = 'm/d/yyyy'

    if 'Price' in headers:
        col_index = headers['Price']
        for row_index in range(2, ws.max_row + 1):
            cell = ws.cell(row_index, col_index)
            if cell.value is not None:
                cell.number_format = '#,##0'

    wb.save(path)


def export_review_queue(detail_df: pd.DataFrame, out_path: Path) -> None:
    overview_df, reason_counts_df, county_counts_df, category_counts_df = build_summary_frames(detail_df)

    with pd.ExcelWriter(out_path, engine='openpyxl', datetime_format='m/d/yyyy') as writer:
        overview_df.to_excel(writer, sheet_name='Summary', index=False, startrow=0, startcol=0)
        reason_counts_df.to_excel(writer, sheet_name='Summary', index=False, startrow=0, startcol=3)
        county_counts_df.to_excel(writer, sheet_name='Summary', index=False, startrow=0, startcol=6)
        category_counts_df.to_excel(writer, sheet_name='Summary', index=False, startrow=0, startcol=9)
        detail_df.to_excel(writer, sheet_name='Review Queue', index=False)

    _style_sheet(out_path, 'Summary', _SUMMARY_WIDTHS, freeze_panes='A2')
    _style_sheet(out_path, 'Review Queue', _DETAIL_WIDTHS, freeze_panes='A2')


def main():
    parser = argparse.ArgumentParser(description='Export the review queue to Excel')
    parser.add_argument('--county', help='Filter review rows by county name')
    parser.add_argument('--reason', action='append',
                        help='Filter by review reason (repeat for multiple reasons)')
    parser.add_argument('--inventory-category', action='append',
                        help='Only export review rows in the given inventory category (repeat for multiple)')
    parser.add_argument('--exclude-inventory-category', action='append',
                        help='Exclude review rows in the given inventory category (repeat for multiple)')
    parser.add_argument('--limit', type=int, help='Limit the number of exported review rows')
    parser.add_argument('--out', help='Output file path (default: OUTPUT_DIR/review_queue.xlsx)')
    args = parser.parse_args()

    sql, params = build_query(
        args.county,
        args.reason,
        args.limit,
        inventory_categories=args.inventory_category,
        exclude_inventory_categories=args.exclude_inventory_category,
    )
    detail_df = fetch_review_rows(sql, params)

    if detail_df.empty:
        print('No review rows matched the filter criteria.')
        return

    out_path = Path(args.out) if args.out else OUTPUT_DIR / 'review_queue.xlsx'
    out_path.parent.mkdir(parents=True, exist_ok=True)

    export_review_queue(detail_df, out_path)
    print(f"Exported {len(detail_df):,} review rows -> {out_path}")


if __name__ == '__main__':
    main()
