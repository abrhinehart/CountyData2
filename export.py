"""
export.py - Export transactions from PostgreSQL to a styled Excel file.

Usage:
    python export.py                                      # All records
    python export.py --county Bay                         # Filter by county
    python export.py --subdivision "PALMETTO COVE"        # Filter by subdivision
    python export.py --from 2023-01-01 --to 2024-01-01   # Date range
    python export.py --exclude-inventory-category scattered_legacy_lots
    python export.py --include-raw                        # Include export_legal_raw column
    python export.py --unmatched-only                     # Only review_flag=TRUE rows
    python export.py --out my_report.xlsx                 # Custom output path
"""

import argparse
from datetime import date
from pathlib import Path

import pandas as pd
import psycopg2
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import DATABASE_URL, OUTPUT_DIR


# Maps DB column names to display headers in the exported Excel file
_COLUMN_MAP = {
    'grantor':       'Grantor',
    'grantee':       'Grantee',
    'type':          'Type',
    'instrument':    'Instrument',
    'date':          'Date',
    'export_legal_desc':    'Export Legal Description',
    'subdivision':   'Subdivision',
    'phase':         'Phase',
    'inventory_category': 'Inventory Category',
    'lots':          'Lots',
    'price':         'Price',
    'price_per_lot': '$ / Lot',
    'acres':         'Acres',
    'price_per_acre':'$ / Acre',
    'county':        'County',
    'notes':         'Notes',
}

# Extra columns available via flags (not in default export)
_EXTRA_COLUMNS = {
    'export_legal_raw': 'Export Legal Raw',
}

_COLUMN_WIDTHS = {
    'Grantor':           50,
    'Grantee':           50,
    'Type':              13,
    'Instrument':        11,
    'Date':              11,
    'Export Legal Description': 50,
    'Subdivision':       30,
    'Phase':              9,
    'Inventory Category': 22,
    'Lots':               5,
    'Price':             10,
    '$ / Lot':            9,
    'Acres':              6,
    '$ / Acre':           8,
    'County':            10,
    'Notes':             40,
    'Export Legal Raw':  80,
}

_CENTER_COLS  = {'Phase', 'Lots', 'Price', '$ / Lot', 'Acres', '$ / Acre', 'County'}
_NUMBER_COLS  = {'Price', '$ / Lot', 'Acres', '$ / Acre'}


def build_query(county: str | None, subdivision: str | None,
                date_from: date | None, date_to: date | None,
                include_raw: bool = False,
                unmatched_only: bool = False,
                inventory_categories: list[str] | None = None,
                exclude_inventory_categories: list[str] | None = None) -> tuple[str, list, dict]:
    col_map = dict(_COLUMN_MAP)
    if include_raw:
        col_map.update(_EXTRA_COLUMNS)

    select_cols = ', '.join(col_map.keys())
    where = []
    params = []

    if county:
        where.append("REPLACE(UPPER(county), ' ', '') = REPLACE(UPPER(%s), ' ', '')")
        params.append(county)
    if subdivision:
        where.append('UPPER(subdivision) = UPPER(%s)')
        params.append(subdivision)
    if date_from:
        where.append('date >= %s')
        params.append(date_from)
    if date_to:
        where.append('date <= %s')
        params.append(date_to)
    if unmatched_only:
        where.append('review_flag = TRUE')
    if inventory_categories:
        where.append('(' + ' OR '.join(['inventory_category = %s'] * len(inventory_categories)) + ')')
        params.extend(inventory_categories)
    if exclude_inventory_categories:
        where.extend(['inventory_category IS DISTINCT FROM %s'] * len(exclude_inventory_categories))
        params.extend(exclude_inventory_categories)

    sql = f"SELECT {select_cols} FROM transactions"
    if where:
        sql += ' WHERE ' + ' AND '.join(where)
    sql += ' ORDER BY date, county, grantor'
    return sql, params, col_map


def fetch_data(sql: str, params: list, col_map: dict) -> pd.DataFrame:
    conn = psycopg2.connect(DATABASE_URL)
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            columns = [desc[0] for desc in cur.description]
            df = pd.DataFrame(cur.fetchall(), columns=columns)
    finally:
        conn.close()
    df.rename(columns=col_map, inplace=True)
    # Ensure all display columns exist (computed columns like $/Lot may be NULL)
    for col in col_map.values():
        if col not in df.columns:
            df[col] = None
    return df[list(col_map.values())]


def apply_styling(path: Path):
    wb = load_workbook(path)
    ws = wb.active
    assert isinstance(ws, Worksheet)

    headers = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1)
               if ws.cell(1, c).value}

    ws.freeze_panes = 'A2'
    for cell in ws[1]:
        cell.font = Font(bold=True)

    if 'Date' in headers:
        ci = headers['Date']
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, ci)
            if cell.value is not None:
                cell.number_format = 'm/d/yyyy'

    for col_name in _NUMBER_COLS:
        if col_name in headers:
            ci = headers[col_name]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, ci)
                if cell.value is not None:
                    cell.number_format = '#,##0'

    for col_name in _CENTER_COLS:
        if col_name in headers:
            ci = headers[col_name]
            for r in range(1, ws.max_row + 1):
                ws.cell(r, ci).alignment = Alignment(horizontal='center')

    for col_name, width in _COLUMN_WIDTHS.items():
        if col_name in headers:
            ws.column_dimensions[get_column_letter(headers[col_name])].width = width

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description='Export transactions to Excel')
    parser.add_argument('--county',       help='Filter by county name (spacing-insensitive)')
    parser.add_argument('--subdivision',  help='Filter by subdivision name')
    parser.add_argument('--from',         dest='date_from', help='Start date (YYYY-MM-DD)')
    parser.add_argument('--to',           dest='date_to',   help='End date (YYYY-MM-DD)')
    parser.add_argument('--include-raw',  action='store_true',
                        help='Include export_legal_raw column in export')
    parser.add_argument('--unmatched-only', action='store_true',
                        help='Only export rows with review_flag=TRUE (unmatched)')
    parser.add_argument('--inventory-category', action='append',
                        help='Only export rows in the given inventory category (repeat for multiple)')
    parser.add_argument('--exclude-inventory-category', action='append',
                        help='Exclude rows in the given inventory category (repeat for multiple)')
    parser.add_argument('--out',          help='Output file path (default: OUTPUT_DIR/export.xlsx)')
    args = parser.parse_args()

    date_from = date.fromisoformat(args.date_from) if args.date_from else None
    date_to   = date.fromisoformat(args.date_to)   if args.date_to   else None

    sql, params, col_map = build_query(
        args.county, args.subdivision, date_from, date_to,
        include_raw=args.include_raw,
        unmatched_only=args.unmatched_only,
        inventory_categories=args.inventory_category,
        exclude_inventory_categories=args.exclude_inventory_category,
    )
    df = fetch_data(sql, params, col_map)

    if df.empty:
        print('No records matched the filter criteria.')
        return

    out_path = Path(args.out) if args.out else OUTPUT_DIR / 'export.xlsx'
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine='openpyxl', datetime_format='m/d/yyyy') as writer:
        df.to_excel(writer, index=False)

    apply_styling(out_path)
    print(f"Exported {len(df):,} records -> {out_path}")


if __name__ == '__main__':
    main()
